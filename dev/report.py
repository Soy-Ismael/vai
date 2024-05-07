import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import openpyxl.chart as charts
from openpyxl.styles import Font
import string
from datetime import datetime
from days import getDay

# Este módulo se creo para realizar reportes de hojas de excel automatizados con las librerias de pandas y openpyxl, dichos reportes seran de los datos que el usuario especifique y de ser posible el reporte tambien se podra realizar segun una inteligencia artificial considere, tambien los reportes pueden incluir graficos, formulas y formatos.

# Para manipular el archivo excel, el mismo debe estar cerrado, no puede ser usado en el momento de la ejecución

ruta_archivo = 'excel/'
nombre_archivo = 'supermarket_sales.xlsx'
nombre_archivo_procesado = 'sales.xlsx'
nombre_hoja = 'Report'

#! Usar parametros en bloque de función
#! Incorporar con IA para análisis de datos y creación de reporte
#? NOTA: Por ahora, el formato del archivo a excel a analizar no debe cambiar, si el archivo de origen cambia, el código debe cambiar para adaptarse a las necesidades de cada empresa.

def create_report(archive_path:str='excel\\', file_name:str = 'supermarket_sales.xlsx', sheet_name:str = 'Report', processed_file:str = 'sales.xlsx'):
    try:
        # 1. Leer archivo excel
        excel_archive = pd.read_excel(archive_path+file_name)
        # Al imprimir excel_archive, obtendre todas las colummnas, para obtener columnas especificar puedo especificar con 2 corchetes
        # print(excel_archive[['Gender', 'Product line', 'Total']])

        # Del archivo de ejemplo queremos saber en que y cuanto gastan las mujeres a comparación de los hombres, para eso podemos utilizar una tabla pivote y se hace de la siguiente manera en python:

        tabla_pivot = excel_archive.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
        # index - Categorias, en este caso mujeres y hombres entonces la categoria es genero
        # columns - Que datos queremos como columnas, en este caso cuanto dinero se destina a cada linea de producto y por eso tomamos product line como columnas
        # values - siempre son datos numericos
        # aggfunc - la función que ejecutaremos, en este caso agregar o sumar y por ultimo que se redonde, no queremos decimales

        # print(tabla_pivot)
        # Exportar el resultado a excel
        tabla_pivot.to_excel(archive_path+processed_file, startrow=4, startcol=1, sheet_name=sheet_name)
        # Recibe como parametros el nombre del archivo, en que fila y columna inicia a exportar respectivamente en el ejemplo pasado, por ultimo recibe el nombre de la hoja


        #* Utilización de openpyxl para añadir graficos y demas herramientas de reportes
        # Leer un archivo con openpyxl

        wb = load_workbook(archive_path+processed_file)
        # recibe como parametro el nombre del archivo y la variable contiene todo el archivo excel con el que openpyxl va a trabajar

        # Especificar que pestaña u hoja queremos manipular con openpyxl
        tab = wb[sheet_name]

        # Para automatizar los graficos debemos automatizar columnas activas e inactivas (columnas minimas y maximas que usamos, donde inicia y termina el reporte), para hacer esto openpyxl nos otorga las siguientes propiedades:
        first_col = wb.active.min_column
        last_col = wb.active.max_column
        first_row = wb.active.min_row
        last_row = wb.active.max_row

        #* Añadir graficas de excel
        # Para hacerlo debemos importar las graficas de openpyxl
        bar_chart = BarChart()

        # Debemos obtener los datos y sus categorias, en este caso, la categoria es hombre/mujer (primera columna) y los datos son todas las demas
        data = Reference(tab, min_col=first_col+1, max_col=last_col, min_row=first_row, max_row=last_row) # Con esto le decimos que los datos estan en la pestaña u hoja que le pasamos como argumento
        # +1 en min_col porque los datos PARA ESTE CASO, estan desde una columna despues de la primera columna, es decir, la primera columna no estara tomada en cuenta como dato

        # Ahora obtengamos la categoria para el grafico
        categoria = Reference(tab, min_col=first_col, max_col=first_col, min_row=first_row+1, max_row=last_row) # Con esto seleccionamos los datos de male y female que son los que utilizaremos como categoria en este caso

        # Ya con los datos referenciados, solo nos falta insertarlos al grafico
        bar_chart.add_data(data, titles_from_data=True) # Esto porque en la data esta incluido el titulo de la tabla
        bar_chart.set_categories(categoria)

        # Para insertar el grafico en el excel:
        tab.add_chart(bar_chart, 'B'+str(last_row+3)) # El segundo parametro corresponde a la posición del grafico, en este caso sera en la columna B, 2 lineas despues de que termine la tabla

        # Añadir estilo al gráfico
        bar_chart.style = 4

        # Añadir titulo del gráfico
        bar_chart.title = 'Ventas'

        # Tambien podemos añadir formular y formates, para ello realizamos cambios antes de guardar...
        # tab['B'+str(last_col+1)] = '=SUMA(B3:B4)'
        #todo tab['C5'] = '=SUMA(C3:C4)' # Esto no funciona actualmente, excel añade un '@' de la nada
        #todo tab['C5'].style = 'Currency' # Estilo moneda, en ingles es 'Currency'

        #* Bucle for
        # Las formulas podemos aplicarlas en tantas celdas como queramos, podemos copiar las formulas una por una o podemos utilizar un bucle for para hacer una función similar a la de arrastrar en excel. Para iniciar a utilizar las letras y no números como referencia utilizaremos un modulo
        abecedario = list(string.ascii_uppercase)
        # print(abecedario)

        abecedario_excel = abecedario[first_col:last_col] # Esto se conoce como slice en listas y me sirve para obtener las vocales desde la primera columna donde hay datos hasta la ultima, las puedo usar según mis necesidades
        # print(abecedario_excel)

        #? comentado por mal funcionamiento por parte de excel (añade un @)
        # for i in abecedario_excel:
        #     tab[f'{i}{last_row+1}'] = f'=SUMA({i}{first_row+1}:{i}{last_row})'
        #     tab[f'{i}{last_row+1}'].style = 'Currency' 
        # tab[f'{abecedario[first_col-1]}{last_row+1}'] = 'Total'

        #* Fuente
        tab['A1'] = 'Reporte'
        tab['A2'] = getDay()

        tab['A1'].font = Font('Arial', b=True, sz=20)
        tab['A2'].font = Font('Abadi', b=False, sz=10)

        # Guardar gráfico
        wb.save(archive_path+processed_file) # Guardamos el archivo con el mismo nombre, esto indica que se guarda el mismo archivo y no se crea otro nuevo
        return True
    except PermissionError as e:
        return 'Por favor, cierre el archivo'


def datos_servicio_report(archive_path:str='excel\\', file_name:str = 'Datos_de _servicio.xlsx', sheet_name:str = 'atenciones', processed_file:str = 'service_repor.xlsx'):
    try:
        # 1. Leer archivo excel
        excel_archive = pd.read_excel(archive_path+file_name)
        # Al imprimir excel_archive, obtendre todas las colummnas, para obtener columnas especificar puedo especificar con 2 corchetes
        # print(excel_archive[['Gender', 'Product line', 'Total']])

        #* Utilización de openpyxl para añadir graficos y demas herramientas de reportes
        # Leer un archivo con openpyxl

        wb = load_workbook(archive_path+file_name)
        # recibe como parametro el nombre del archivo y la variable contiene todo el archivo excel con el que openpyxl va a trabajar

        # Especificar que pestaña u hoja queremos manipular con openpyxl
        tab = wb[sheet_name]

        # Para automatizar los graficos debemos automatizar columnas activas e inactivas (columnas minimas y maximas que usamos, donde inicia y termina el reporte), para hacer esto openpyxl nos otorga las siguientes propiedades:
        first_col = wb.active.min_column
        last_col = wb.active.max_column
        first_row = wb.active.min_row
        last_row = wb.active.max_row

        print(first_col)
        print(last_col)
        print(first_row)
        print(last_row)

        #* Añadir graficas de excel
        # Para hacerlo debemos importar las graficas de openpyxl
        bar_chart = BarChart()
        bubble_chart = charts.BubbleChart()

        # Debemos obtener los datos y sus categorias, en este caso, la categoria es hombre/mujer (primera columna) y los datos son todas las demas
        data = Reference(tab, min_col=first_col+1, max_col=last_col, min_row=first_row, max_row=last_row) # Con esto le decimos que los datos estan en la pestaña u hoja que le pasamos como argumento
        # +1 en min_col porque los datos PARA ESTE CASO, estan desde una columna despues de la primera columna, es decir, la primera columna no estara tomada en cuenta como dato

        # Ahora obtengamos la categoria para el grafico
        categoria = Reference(tab, min_col=first_col, max_col=first_col, min_row=first_row+1, max_row=last_row) # Con esto seleccionamos los datos de male y female que son los que utilizaremos como categoria en este caso

        # Ya con los datos referenciados, solo nos falta insertarlos al grafico
        bar_chart.add_data(data, titles_from_data=True) # Esto porque en la data esta incluido el titulo de la tabla
        bar_chart.set_categories(categoria)

        bubble_chart.add_data(data, titles_from_data=True)
        bubble_chart.set_categories(categoria)

        # Para insertar el grafico en el excel:
        tab.add_chart(bar_chart, 'B'+str(last_row+3)) # El segundo parametro corresponde a la posición del grafico, en este caso sera en la columna B, 2 lineas despues de que termine la tabla
        
        print(bar_chart.layout)
        tab.add_chart(bubble_chart, 'B'+str(last_row+15))
        # Añadir estilo al gráfico
        bar_chart.style = 4
        bubble_chart.style = 5

        # Añadir titulo del gráfico
        bar_chart.title = 'Ventas'
        bubble_chart.title = 'Ventas Bubble'

        # Tambien podemos añadir formular y formates, para ello realizamos cambios antes de guardar...
        # tab['B'+str(last_col+1)] = '=SUMA(B3:B4)'
        #todo tab['C5'] = '=SUMA(C3:C4)' # Esto no funciona actualmente, excel añade un '@' de la nada
        #todo tab['C5'].style = 'Currency' # Estilo moneda, en ingles es 'Currency'

        #* Bucle for
        # Las formulas podemos aplicarlas en tantas celdas como queramos, podemos copiar las formulas una por una o podemos utilizar un bucle for para hacer una función similar a la de arrastrar en excel. Para iniciar a utilizar las letras y no números como referencia utilizaremos un modulo
        abecedario = list(string.ascii_uppercase)
        # print(abecedario)

        abecedario_excel = abecedario[first_col:last_col] # Esto se conoce como slice en listas y me sirve para obtener las vocales desde la primera columna donde hay datos hasta la ultima, las puedo usar según mis necesidades
        # print(abecedario_excel)

        #? comentado por mal funcionamiento por parte de excel (añade un @)
        # for i in abecedario_excel:
        #     tab[f'{i}{last_row+1}'] = f'=SUMA({i}{first_row+1}:{i}{last_row})'
        #     tab[f'{i}{last_row+1}'].style = 'Currency' 
        # tab[f'{abecedario[first_col-1]}{last_row+1}'] = 'Total'

        #* Fuente
        tab['A1'] = 'Reporte'
        tab['A2'] = getDay()

        tab['A1'].font = Font('Arial', b=True, sz=20)
        tab['A2'].font = Font('Abadi', b=False, sz=10)

        # Guardar gráfico
        wb.save(archive_path+processed_file) # Guardamos el archivo con el mismo nombre, esto indica que se guarda el mismo archivo y no se crea otro nuevo
        return True
    except PermissionError as e:
        return 'Por favor, cierre el archivo'


result = datos_servicio_report()
# result = create_report()
# print(result)