import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import openpyxl.chart as charts
from openpyxl.styles import Font
import string
from datetime import datetime
from days import getDay

# Este módulo se creo para realizar reportes de hojas de excel automatizados con las librerias de pandas y openpyxl, dichos reportes seran de los datos que el usuario especifique y de ser posible el reporte tambien se podra realizar segun una inteligencia artificial considere, tambien los reportes pueden incluir graficos, formulas y formatos.

# Para manipular el archivo excel, el mismo debe estar cerrado, no puede ser usado en el momento de la ejecución

ruta_archivo = 'excel/'
nombre_archivo = 'supermarket_sales.xlsx'
nombre_archivo_procesado = 'sales.xlsx'
nombre_hoja = 'Report'

#! Usar parametros en bloque de función
#! Incorporar con IA para análisis de datos y creación de reporte
#? NOTA: Por ahora, el formato del archivo a excel a analizar no debe cambiar, si el archivo de origen cambia, el código debe cambiar para adaptarse a las necesidades de cada empresa.

def create_report(archive_path:str='excel\\', file_name:str = 'supermarket_sales.xlsx', sheet_name:str = 'Report', processed_file:str = 'sales.xlsx'):
    try:
        # 1. Leer archivo excel
        excel_archive = pd.read_excel(archive_path+file_name)
        tabla_pivot = excel_archive.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
        tabla_pivot.to_excel(archive_path+processed_file, startrow=4, startcol=1, sheet_name=sheet_name)

        #* Utilización de openpyxl para añadir graficos y demas herramientas de reportes

        wb = load_workbook(archive_path+processed_file)
        tab = wb[sheet_name]

        first_col = wb.active.min_column
        last_col = wb.active.max_column
        first_row = wb.active.min_row
        last_row = wb.active.max_row

        #* Añadir graficas de excel
        bar_chart = BarChart()

        data = Reference(tab, min_col=first_col+1, max_col=last_col, min_row=first_row, max_row=last_row)
        categoria = Reference(tab, min_col=first_col, max_col=first_col, min_row=first_row+1, max_row=last_row)

        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categoria)

        tab.add_chart(bar_chart, 'B'+str(last_row+3))
        bar_chart.style = 4
        bar_chart.title = 'Ventas'

        abecedario = list(string.ascii_uppercase)
        abecedario_excel = abecedario[first_col:last_col]

        #* Fuente
        tab['A1'] = 'Reporte'
        tab['A2'] = getDay()

        tab['A1'].font = Font('Arial', b=True, sz=20)
        tab['A2'].font = Font('Abadi', b=False, sz=10)

        # Guardar gráfico
        wb.save(archive_path+processed_file)
        return True
    except PermissionError as e:
        return 'Por favor, cierre el archivo'


# result = datos_servicio_report()
# result = create_report()
# print(result)